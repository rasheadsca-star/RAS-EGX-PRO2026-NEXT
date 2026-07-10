#!/usr/bin/env python3
"""Import a reviewed historical XLSX seed without overwriting valid live history.

Safety properties:
- Imports only tickers already present in data/symbol-map.json.
- Rejects ISIN conflicts when both the map and workbook provide ISIN.
- Rejects invalid OHLC rows; never estimates or repairs prices.
- Keeps existing rows on duplicate dates and skips existing complete 100-session files.
- Writes atomically and records every imported, skipped, rejected, and quarantined item.
"""

from __future__ import annotations

import json
import math
import os
import re
import sys
import tempfile
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

ROOT = Path(os.environ.get("GITHUB_WORKSPACE", Path.cwd())).resolve()
DATA = ROOT / "data"
CONFIG_PATH = DATA / "history-seed-import-config.json"
MAP_PATH = DATA / "symbol-map.json"
SUMMARY_PATH = DATA / "history-summary.json"
HISTORY_DIR = DATA / "history"
REPORT_PATH = DATA / "history-seed-import-report.json"
QUARANTINE_PATH = DATA / "history-seed-quarantine.json"
LAST_RUN_PATH = DATA / "history-last-run.json"
STATE_PATH = DATA / "history-batch-state.json"
SOURCE_AUDIT_PATH = DATA / "source-audit.json"
FALLBACK_QUEUE_PATH = DATA / "history-fallback-queue.json"

REQUIRED_HEADERS = {
    "الرمز": "ticker",
    "اسم الشركة": "company_name",
    "ISIN": "isin",
    "التاريخ": "session_date",
    "الفتح": "open",
    "الأعلى": "high",
    "الأدنى": "low",
    "الإغلاق": "close",
    "الحجم": "volume",
    "حالة الجودة": "quality_status",
    "الثقة": "seed_confidence",
    "سبب الملاحظة": "note",
    "المصدر": "source_label",
    "رابط المصدر": "source_url",
}


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def read_json(path: Path, default: Any = None) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        return default


def write_json_atomic(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=str(path.parent))
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
        os.replace(tmp_name, path)
    finally:
        if os.path.exists(tmp_name):
            os.unlink(tmp_name)


def safe_ticker(value: Any) -> str:
    ticker = re.sub(r"[^A-Z0-9]", "", str(value or "").strip().upper())
    return ticker[:16]


def normalize_isin(value: Any) -> Optional[str]:
    text = re.sub(r"\s+", "", str(value or "").upper())
    return text if re.fullmatch(r"[A-Z]{2}[A-Z0-9]{9}[0-9]", text) else None


def parse_date(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def integer_or_none(value: Any) -> Optional[int]:
    parsed = number(value)
    if parsed is None:
        return None
    if parsed < 0:
        return None
    return int(round(parsed))


def validate_ohlc(row: Dict[str, Any]) -> List[str]:
    errors: List[str] = []
    session_date = row.get("date")
    if not session_date:
        errors.append("date_invalid")
    elif session_date > date.today().isoformat():
        errors.append("date_in_future")

    o, h, l, c = row.get("open"), row.get("high"), row.get("low"), row.get("close")
    for field, value in (("open", o), ("high", h), ("low", l), ("close", c)):
        if value is None:
            errors.append(f"{field}_missing")
        elif value <= 0:
            errors.append(f"{field}_not_positive")

    if all(value is not None for value in (o, h, l, c)):
        if h < l:
            errors.append("high_below_low")
        if h < o:
            errors.append("high_below_open")
        if h < c:
            errors.append("high_below_close")
        if l > o:
            errors.append("low_above_open")
        if l > c:
            errors.append("low_above_close")

    volume = row.get("volume")
    if volume is not None and volume < 0:
        errors.append("volume_negative")
    return errors


def normalize_map(raw: Any) -> Tuple[Dict[str, Dict[str, Any]], bool]:
    was_array = isinstance(raw, list)
    entries = raw if was_array else list((raw or {}).values())
    result: Dict[str, Dict[str, Any]] = {}
    for item in entries:
        if not isinstance(item, dict):
            continue
        ticker = safe_ticker(item.get("ticker"))
        if ticker:
            result[ticker] = {**item, "ticker": ticker}
    return result, was_array


def serialize_map(entries: Dict[str, Dict[str, Any]], was_array: bool) -> Any:
    ordered = [entries[key] for key in sorted(entries)]
    if was_array:
        return ordered
    return {entry["ticker"]: entry for entry in ordered}


def load_existing_history(ticker: str) -> Optional[Dict[str, Any]]:
    return read_json(HISTORY_DIR / f"{ticker}.json", None)


def valid_existing_sessions(document: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
    sessions = document.get("sessions", []) if isinstance(document, dict) else []
    output: List[Dict[str, Any]] = []
    seen = set()
    for session in sessions if isinstance(sessions, list) else []:
        if not isinstance(session, dict):
            continue
        normalized = {
            **session,
            "date": parse_date(session.get("date")),
            "open": number(session.get("open")),
            "high": number(session.get("high")),
            "low": number(session.get("low")),
            "close": number(session.get("close")),
            "volume": integer_or_none(session.get("volume")),
        }
        if validate_ohlc(normalized) or not normalized["date"] or normalized["date"] in seen:
            continue
        seen.add(normalized["date"])
        output.append(normalized)
    return sorted(output, key=lambda item: item["date"])


def session_confidence(session: Dict[str, Any]) -> float:
    confidence = session.get("confidence")
    if isinstance(confidence, dict):
        return float(confidence.get("overall") or 0)
    try:
        return float(confidence or 0)
    except (TypeError, ValueError):
        return 0.0


def history_status(count: int) -> str:
    if count >= 100:
        return "historical_complete_100"
    if count >= 50:
        return "historical_partial_50"
    if count >= 20:
        return "historical_limited_20"
    if count >= 5:
        return "historical_limited_5"
    return "historical_insufficient"


def unique(values: Iterable[Any]) -> List[Any]:
    output = []
    seen = set()
    for value in values:
        if value in (None, ""):
            continue
        key = json.dumps(value, ensure_ascii=False, sort_keys=True) if isinstance(value, (dict, list)) else str(value)
        if key not in seen:
            seen.add(key)
            output.append(value)
    return output


def build_session(
    ticker: str,
    row: Dict[str, Any],
    confidence: int,
    identity_confidence: int,
    config: Dict[str, Any],
    imported_at: str,
) -> Dict[str, Any]:
    warnings = ["historical_seed_single_source", "historical_seed_requires_recent_gap_fill"]
    if row.get("volume") is None:
        warnings.append("volume_missing")
    note = str(row.get("note") or "").strip()
    if note:
        warnings.append(f"seed_note:{note[:160]}")
    return {
        "ticker": ticker,
        "date": row["date"],
        "open": row["open"],
        "high": row["high"],
        "low": row["low"],
        "close": row["close"],
        "adjustedClose": None,
        "volume": row.get("volume"),
        "currency": "EGP",
        "primarySource": config["sourceKey"],
        "officialVerified": False,
        "verifiedBy": [],
        "sourceUrls": {"primary": config.get("sourceUrl"), "verification": []},
        "fetchedAt": imported_at,
        "validatedAt": imported_at,
        "confidence": {
            "overall": confidence,
            "ohlc": confidence,
            "volume": 60 if row.get("volume") is None else confidence,
            "symbolIdentity": identity_confidence,
        },
        "validationStatus": "historical_seed_single_source_validated",
        "warnings": unique(warnings),
    }


def merge_sessions(seed: List[Dict[str, Any]], existing: List[Dict[str, Any]], limit: int) -> List[Dict[str, Any]]:
    # Seed first; current repository history always wins on duplicate dates.
    by_date: Dict[str, Dict[str, Any]] = {item["date"]: item for item in seed}
    for item in existing:
        by_date[item["date"]] = item
    return [by_date[key] for key in sorted(by_date)[-limit:]]


def update_source_audit(report: Dict[str, Any]) -> None:
    raw = read_json(SOURCE_AUDIT_PATH, [])
    record = {
        "operation": "historical_seed_xlsx_import",
        "startedAt": report["startedAt"],
        "completedAt": report["completedAt"],
        "source": report["source"],
        "workbook": report["workbook"],
        "rowsRead": report["rowsRead"],
        "rowsValid": report["rowsValid"],
        "rowsQuarantined": report["rowsQuarantined"],
        "symbolsImportedOrImproved": report["counts"]["importedOrImproved"],
        "warnings": ["single_source_seed", "latest_seed_session_is_historical"],
    }
    if isinstance(raw, list):
        raw.append(record)
        raw = raw[-500:]
    elif isinstance(raw, dict):
        operations = raw.get("operations") if isinstance(raw.get("operations"), list) else []
        operations.append(record)
        raw["operations"] = operations[-500:]
        raw["lastOperation"] = record
    else:
        raw = [record]
    write_json_atomic(SOURCE_AUDIT_PATH, raw)


def main() -> int:
    started_at = now_iso()
    config = read_json(CONFIG_PATH, None)
    if not isinstance(config, dict) or config.get("enabled") is not True:
        raise RuntimeError("Seed importer is disabled or config is missing")

    workbook_path = ROOT / config.get("workbookPath", "")
    if not workbook_path.is_file():
        raise RuntimeError(f"Workbook not found: {workbook_path.relative_to(ROOT)}")

    raw_map = read_json(MAP_PATH, None)
    if raw_map is None:
        raise RuntimeError("Missing data/symbol-map.json")
    symbol_map, map_was_array = normalize_map(raw_map)
    if not symbol_map:
        raise RuntimeError("Symbol map is empty")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if config["sheetName"] not in workbook.sheetnames:
        raise RuntimeError(f"Sheet not found: {config['sheetName']}")
    sheet = workbook[config["sheetName"]]
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise RuntimeError("Workbook sheet is empty")
    header_index = {str(value).strip(): index for index, value in enumerate(header_row) if value not in (None, "")}
    missing_headers = [header for header in REQUIRED_HEADERS if header not in header_index]
    if missing_headers:
        raise RuntimeError(f"Missing required headers: {', '.join(missing_headers)}")

    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    ticker_identity: Dict[str, Dict[str, Any]] = {}
    quarantine: List[Dict[str, Any]] = []
    rows_read = 0
    rows_valid = 0

    for excel_row_number, values in enumerate(rows, start=2):
        rows_read += 1
        get = lambda header: values[header_index[header]] if header_index[header] < len(values) else None
        ticker = safe_ticker(get("الرمز"))
        parsed = {
            "ticker": ticker,
            "company_name": str(get("اسم الشركة") or "").strip() or None,
            "isin": normalize_isin(get("ISIN")),
            "date": parse_date(get("التاريخ")),
            "open": number(get("الفتح")),
            "high": number(get("الأعلى")),
            "low": number(get("الأدنى")),
            "close": number(get("الإغلاق")),
            "volume": integer_or_none(get("الحجم")),
            "quality_status": str(get("حالة الجودة") or "").strip() or None,
            "note": str(get("سبب الملاحظة") or "").strip() or None,
        }
        errors = []
        if not ticker:
            errors.append("ticker_invalid")
        elif config.get("requireTickerInSymbolMap", True) and ticker not in symbol_map:
            errors.append("ticker_not_in_symbol_map")
        errors.extend(validate_ohlc(parsed))
        if errors:
            quarantine.append({"row": excel_row_number, "ticker": ticker or None, "date": parsed.get("date"), "errors": unique(errors)})
            continue
        grouped[ticker].append(parsed)
        rows_valid += 1
        ticker_identity.setdefault(ticker, {"isin": parsed["isin"], "companyName": parsed["company_name"]})

    imported = []
    improved = []
    skipped_complete = []
    skipped_no_improvement = []
    rejected_symbols = []
    max_sessions = int(config.get("maximumStoredSessions", 100))
    imported_at = now_iso()

    for ticker in sorted(grouped):
        entry = symbol_map.get(ticker)
        if not entry or entry.get("active") is False:
            rejected_symbols.append({"ticker": ticker, "reason": "ticker_inactive_or_missing"})
            continue
        existing_doc = load_existing_history(ticker)
        existing_sessions = valid_existing_sessions(existing_doc)
        if config.get("skipExistingComplete100", True) and len(existing_sessions) >= 100:
            skipped_complete.append(ticker)
            continue

        book_isin = ticker_identity.get(ticker, {}).get("isin")
        map_isin = normalize_isin(entry.get("isin"))
        if config.get("rejectIsinMismatch", True) and map_isin and book_isin and map_isin != book_isin:
            rejected_symbols.append({"ticker": ticker, "reason": "isin_mismatch", "mapIsin": map_isin, "workbookIsin": book_isin})
            continue

        isin_match = bool(map_isin and book_isin and map_isin == book_isin)
        confidence = int(config.get("confidenceWithIsinMatch" if isin_match else "confidenceWithTickerMapMatch", 75 if isin_match else 65))
        identity_confidence = 100 if isin_match else 80
        deduped_rows: Dict[str, Dict[str, Any]] = {}
        for row in grouped[ticker]:
            deduped_rows[row["date"]] = row
        seed_sessions = [
            build_session(ticker, deduped_rows[key], confidence, identity_confidence, config, imported_at)
            for key in sorted(deduped_rows)
        ]
        merged = merge_sessions(seed_sessions, existing_sessions, max_sessions)
        minimum = int(config.get("minimumValidSessionsForNewFile", 20)) if not existing_sessions else 5
        if len(merged) < minimum:
            rejected_symbols.append({"ticker": ticker, "reason": "insufficient_valid_sessions", "validSessions": len(merged), "minimum": minimum})
            continue
        improvement = len(merged) - len(existing_sessions)
        if improvement < int(config.get("minimumImprovementSessions", 1)) and (not merged or not existing_sessions or merged[-1]["date"] <= existing_sessions[-1]["date"]):
            skipped_no_improvement.append(ticker)
            continue

        latest = merged[-1]["date"]
        seed_only = not existing_sessions
        average_confidence = round(sum(session_confidence(item) for item in merged) / len(merged), 2)
        warnings = list(existing_doc.get("warnings", [])) if isinstance(existing_doc, dict) else []
        warnings.extend(["historical_seed_requires_recent_gap_fill", "historical_seed_not_officially_verified"])
        quarantined_for_ticker = sum(1 for item in quarantine if item.get("ticker") == ticker)
        if quarantined_for_ticker:
            warnings.append(f"historical_seed_rows_quarantined:{quarantined_for_ticker}")
        primary_source = config["sourceKey"] if seed_only else (existing_doc.get("primarySource") or "mixed_history")
        if existing_sessions and seed_sessions:
            primary_source = "mixed_history_with_kaggle_yahoo_seed"

        document = {
            "schemaVersion": "12.6.0",
            "ticker": ticker,
            "companyNameAr": entry.get("companyNameAr"),
            "companyNameEn": entry.get("companyNameEn") or ticker_identity.get(ticker, {}).get("companyName"),
            "isin": map_isin or book_isin,
            "reutersCode": entry.get("reutersCode"),
            "yahooSymbol": entry.get("yahooSymbol"),
            "currency": entry.get("currency") or "EGP",
            "exchange": entry.get("exchange") or "EGX",
            "generatedAt": imported_at,
            "availableSessions": len(merged),
            "firstSession": merged[0]["date"],
            "lastSession": latest,
            "historyStatus": history_status(len(merged)),
            "primarySource": primary_source,
            "verificationSources": unique(existing_doc.get("verificationSources", []) if isinstance(existing_doc, dict) else []),
            "officiallyVerifiedLatestSession": bool(existing_doc.get("officiallyVerifiedLatestSession")) if isinstance(existing_doc, dict) and latest == existing_doc.get("lastSession") else False,
            "symbolVerified": True,
            "symbolVerification": {
                "verified": True,
                "policy": "historical_seed_isin_match" if isin_match else "historical_seed_ticker_map_match",
                "identityConfidence": identity_confidence,
                "priceOfficiallyVerified": False,
                "source": config["sourceKey"],
            },
            "averageConfidence": average_confidence,
            "staleData": bool(config.get("markSeedOnlyFilesStale", True) and latest <= config.get("sourceLatestSession", latest)),
            "updateFailed": False,
            "warnings": unique(warnings),
            "seedImport": {
                "source": config["sourceKey"],
                "sourceUrl": config.get("sourceUrl"),
                "importedAt": imported_at,
                "seedValidSessions": len(seed_sessions),
                "seedQuarantinedRows": quarantined_for_ticker,
                "seedLatestSession": seed_sessions[-1]["date"] if seed_sessions else None,
                "requiresIncrementalGapFill": True,
            },
            "sessions": merged,
        }
        write_json_atomic(HISTORY_DIR / f"{ticker}.json", document)
        if not map_isin and book_isin:
            entry["isin"] = book_isin
        entry.update({
            "seedHistoryAvailable": True,
            "seedHistoryImportedAt": imported_at,
            "seedHistoryValidSessions": len(seed_sessions),
            "seedHistoryLastSession": seed_sessions[-1]["date"] if seed_sessions else None,
            "seedHistorySource": config["sourceKey"],
            "fallbackRequired": len(merged) < 100,
        })
        symbol_map[ticker] = entry
        record = {"ticker": ticker, "before": len(existing_sessions), "after": len(merged), "added": max(0, improvement), "lastSession": latest, "confidence": average_confidence}
        (improved if existing_sessions else imported).append(record)

    write_json_atomic(MAP_PATH, serialize_map(symbol_map, map_was_array))

    # Move symbols with usable imported history out of the unresolved-source queue.
    resolved_records = imported + improved
    resolved_tickers = {item["ticker"] for item in resolved_records if int(item.get("after", 0)) >= 20}
    fallback_queue = read_json(FALLBACK_QUEUE_PATH, None)
    if isinstance(fallback_queue, dict) and isinstance(fallback_queue.get("queue"), list) and resolved_tickers:
        remaining = [item for item in fallback_queue["queue"] if safe_ticker(item.get("ticker")) not in resolved_tickers]
        prior_resolved = fallback_queue.get("resolvedByHistoricalSeed") if isinstance(fallback_queue.get("resolvedByHistoricalSeed"), list) else []
        prior_by_ticker = {safe_ticker(item.get("ticker")): item for item in prior_resolved if safe_ticker(item.get("ticker"))}
        for item in resolved_records:
            if item["ticker"] not in resolved_tickers:
                continue
            prior_by_ticker[item["ticker"]] = {
                "ticker": item["ticker"],
                "resolvedAt": imported_at,
                "resolvedBy": config["sourceKey"],
                "availableSessions": item["after"],
                "lastSession": item["lastSession"],
                "stillRequiresRecentGapFill": True,
            }
        fallback_queue["schemaVersion"] = "12.6.0"
        fallback_queue["generatedAt"] = imported_at
        fallback_queue["queue"] = remaining
        fallback_queue["total"] = len(remaining)
        fallback_queue["resolvedByHistoricalSeed"] = [prior_by_ticker[key] for key in sorted(prior_by_ticker)]
        write_json_atomic(FALLBACK_QUEUE_PATH, fallback_queue)

    completed_at = now_iso()
    report = {
        "schemaVersion": "12.6.0",
        "startedAt": started_at,
        "completedAt": completed_at,
        "mode": "import_seed_xlsx",
        "workbook": str(workbook_path.relative_to(ROOT)),
        "sheet": config["sheetName"],
        "source": {"key": config["sourceKey"], "label": config.get("sourceLabel"), "url": config.get("sourceUrl"), "latestSession": config.get("sourceLatestSession"), "official": False},
        "rowsRead": rows_read,
        "rowsValid": rows_valid,
        "rowsQuarantined": len(quarantine),
        "symbolsInWorkbook": len(grouped),
        "counts": {
            "newFilesImported": len(imported),
            "existingFilesImproved": len(improved),
            "importedOrImproved": len(imported) + len(improved),
            "skippedExistingComplete100": len(skipped_complete),
            "skippedNoImprovement": len(skipped_no_improvement),
            "rejectedSymbols": len(rejected_symbols),
        },
        "imported": imported,
        "improved": improved,
        "skippedExistingComplete100": skipped_complete,
        "skippedNoImprovement": skipped_no_improvement,
        "rejectedSymbols": rejected_symbols,
        "resolvedFallbackTickers": sorted(resolved_tickers),
        "warnings": [
            "Seed is a historical single-source dataset and is not official EGX verification.",
            "Imported files remain stale until the post-2026-02-04 gap is filled by a current source.",
            "Invalid OHLC rows were quarantined and never repaired or estimated.",
        ],
    }
    write_json_atomic(REPORT_PATH, report)
    write_json_atomic(QUARANTINE_PATH, {"schemaVersion": "12.6.0", "generatedAt": completed_at, "count": len(quarantine), "rows": quarantine})
    update_source_audit(report)

    previous_last_run = read_json(LAST_RUN_PATH, {})
    write_json_atomic(LAST_RUN_PATH, {
        **(previous_last_run if isinstance(previous_last_run, dict) else {}),
        "schemaVersion": "12.6.0",
        "generatedAt": completed_at,
        "mode": "import_seed_xlsx",
        "selectedTickers": sorted(grouped),
        "succeededTickers": [item["ticker"] for item in imported + improved],
        "skippedCompleteTickers": skipped_complete,
        "failed": rejected_symbols,
        "seedImportReport": "data/history-seed-import-report.json",
    })
    previous_state = read_json(STATE_PATH, {})
    write_json_atomic(STATE_PATH, {
        **(previous_state if isinstance(previous_state, dict) else {}),
        "schemaVersion": "12.6.0",
        "generatedAt": completed_at,
        "lastMode": "import_seed_xlsx",
        "lastSelectedCount": len(grouped),
        "lastFetchedCount": 0,
        "lastSkippedCompleteCount": len(skipped_complete),
        "lastSucceededCount": len(imported) + len(improved),
        "lastFailedCount": len(rejected_symbols),
    })

    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if (imported or improved or skipped_complete) else 1


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:  # fail closed with a clear diagnostic
        print(f"V12.6 seed import failed: {exc}", file=sys.stderr)
        sys.exit(1)
